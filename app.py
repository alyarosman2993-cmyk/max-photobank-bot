import os
import re
import requests
import certifi
import ssl

ssl._create_default_https_context = lambda: ssl.create_default_context(cafile=certifi.where())
from io import BytesIO
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from flask import Flask, request, jsonify

app = Flask(__name__)

MAX_TOKEN = os.environ.get("MAX_TOKEN")
MAX_API = "https://platform-api2.max.ru"
YANDEX_EXCEL_PATH = os.environ.get("YANDEX_EXCEL_PATH")

USER_STATES = {}

PARTICIPANT_TYPES = [
    "Субъект РФ",
    "ФОИВ",
    "ВУЗ",
    "СУЗ",
    "Школа",
    "Организации и НКО",
    "Политические НКО",
]

ACTIONS = [
    "День семьи, любви и верности",
    "День физкультурника",
    "День Государственного флага РФ",
    "День воссоединения исторических регионов",
    "День рождения Президента РФ",
    "День отца",
    "День народного единства",
    "День матери",
    "День Героев Отечества",
    "Новый год",
]

SHEET_BY_TYPE = {
    "Субъект РФ": "Субъекты РФ",
    "ФОИВ": "ФОИВ",
    "ВУЗ": "ВУЗы",
    "СУЗ": "СУЗы",
    "Школа": "Школы",
    "Организации и НКО": "Организации и НКО",
    "Политические НКО": "Политические НКО",
}


@app.route("/")
def home():
    return "MAX Photobank Bot работает"


def make_keyboard(items):
    return [{
        "type": "inline_keyboard",
        "payload": {
            "buttons": [[{"type": "message", "text": item}] for item in items]
        }
    }]


def send_message(chat_id, text, buttons=None):
    url = f"{MAX_API}/messages?chat_id={chat_id}"
    headers = {
        "Authorization": MAX_TOKEN,
        "Content-Type": "application/json"
    }
    payload = {"text": text}

    if buttons:
        payload["attachments"] = make_keyboard(buttons)

    response = requests.post(url, json=payload, headers=headers)
    print("SEND RESPONSE:", response.status_code, response.text, flush=True)
    return response


def yandex_headers():
    return {
        "Authorization": f"OAuth {os.environ.get('YANDEX_TOKEN')}"
    }


def clean_broken_excel_tables(xlsx_bytes):
    input_file = BytesIO(xlsx_bytes)
    output_file = BytesIO()

    with ZipFile(input_file, "r") as zin:
        with ZipFile(output_file, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename

                if name.startswith("xl/tables/"):
                    continue

                data = zin.read(name)

                if name == "[Content_Types].xml":
                    root = ET.fromstring(data)
                    for child in list(root):
                        if child.attrib.get("PartName", "").startswith("/xl/tables/"):
                            root.remove(child)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                if name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
                    root = ET.fromstring(data)
                    for child in list(root):
                        if child.attrib.get("Type", "").endswith("/table"):
                            root.remove(child)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    root = ET.fromstring(data)
                    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                    for table_parts in root.findall("main:tableParts", ns):
                        root.remove(table_parts)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                zout.writestr(item, data)

    output_file.seek(0)
    return output_file


def download_excel_from_yandex():
    response = requests.get(
        "https://cloud-api.yandex.net/v1/disk/resources/download",
        headers=yandex_headers(),
        params={"path": YANDEX_EXCEL_PATH}
    )
    response.raise_for_status()

    download_url = response.json()["href"]
    file_response = requests.get(download_url)
    file_response.raise_for_status()

    return clean_broken_excel_tables(file_response.content)


def upload_excel_to_yandex(file_bytes):
    response = requests.get(
        "https://cloud-api.yandex.net/v1/disk/resources/upload",
        headers=yandex_headers(),
        params={"path": YANDEX_EXCEL_PATH, "overwrite": "true"}
    )
    response.raise_for_status()

    upload_url = response.json()["href"]
    upload_response = requests.put(upload_url, data=file_bytes.getvalue())
    upload_response.raise_for_status()


def add_row_to_first_empty(ws, row):
    target_row = 2

    while ws.cell(row=target_row, column=1).value:
        target_row += 1

    for col, value in enumerate(row, start=1):
        ws.cell(row=target_row, column=col, value=value)


def save_application_to_excel(state):
    file_stream = download_excel_from_yandex()
    workbook = load_workbook(file_stream)

    row = [
        datetime.now().strftime("%d.%m.%Y %H:%M"),
        state.get("participant_type", ""),
        state.get("participant_name", ""),
        state.get("region", ""),
        state.get("action", ""),
        state.get("format", ""),
        state.get("disk_link", ""),
    ]

    main_sheet = workbook["Все заявки"]
    add_row_to_first_empty(main_sheet, row)

    profile_sheet_name = SHEET_BY_TYPE.get(state.get("participant_type"))
    if profile_sheet_name and profile_sheet_name in workbook.sheetnames:
        profile_sheet = workbook[profile_sheet_name]
        add_row_to_first_empty(profile_sheet, row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    upload_excel_to_yandex(output)


@app.route("/check-yandex")
def check_yandex():
    response = requests.get(
        "https://cloud-api.yandex.net/v1/disk/resources",
        headers=yandex_headers(),
        params={"path": YANDEX_EXCEL_PATH}
    )

    return jsonify({
        "status_code": response.status_code,
        "response": response.text
    })


def start_screen(chat_id):
    send_message(
        chat_id,
        "Здравствуйте!\n\n"
        "Данный бот создан для сбора фото- и видеоматериалов по реализации "
        "Всероссийских акций, приуроченных к государственным праздникам Российской Федерации.\n\n"
        "Для направления материалов нажмите кнопку ниже.",
        ["Начать работу"]
    )


def start_flow(chat_id, user_id):
    USER_STATES[user_id] = {"step": "participant_type"}
    send_message(
        chat_id,
        "Для направления материалов Вам потребуется последовательно указать:\n"
        "• тип участника;\n"
        "• наименование участника — кроме случая «Субъект РФ»;\n"
        "• регион проведения;\n"
        "• Всероссийскую акцию;\n"
        "• формат мероприятия;\n"
        "• ссылку на Яндекс.Диск с фото- и видеоматериалами.\n\n"
        "Выберите тип участника:",
        PARTICIPANT_TYPES
    )


def get_message_parts(data):
    message = data.get("message", {})
    body = message.get("body", {})

    chat_id = message.get("recipient", {}).get("chat_id")
    user_id = message.get("sender", {}).get("user_id")
    text = body.get("text", "")

    text = text.strip() if text else ""
    return chat_id, user_id, text


def has_link(text):
    return bool(re.search(r"https?://\S+", text or ""))


@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json
    print("MAX EVENT:", data, flush=True)

    try:
        chat_id, user_id, text = get_message_parts(data)

        if not chat_id or not user_id:
            return jsonify({"status": "ok"})

        state = USER_STATES.get(user_id)

        if text in ["/start", "старт", "Старт", "начать", "Начать"]:
            start_screen(chat_id)
            return jsonify({"status": "ok"})

        if text == "Начать работу":
            start_flow(chat_id, user_id)
            return jsonify({"status": "ok"})

        if text == "Добавить ещё один формат":
            old_state = USER_STATES.get(user_id, {})
            USER_STATES[user_id] = {
                "step": "format",
                "participant_type": old_state.get("participant_type"),
                "participant_name": old_state.get("participant_name"),
                "region": old_state.get("region"),
                "action": old_state.get("action")
            }
            send_message(chat_id, "Введите следующий формат мероприятия:")
            return jsonify({"status": "ok"})

        if text == "Завершить работу":
            USER_STATES.pop(user_id, None)
            send_message(
                chat_id,
                "Благодарим за направление материалов!\n\n"
                "Для повторной загрузки материалов Вы можете вернуться в бот "
                "в любое время и нажать кнопку «Начать работу».",
                ["Начать работу"]
            )
            return jsonify({"status": "ok"})

        if not state:
            start_screen(chat_id)
            return jsonify({"status": "ok"})

        if state["step"] == "participant_type":
            if text not in PARTICIPANT_TYPES:
                send_message(chat_id, "Пожалуйста, выберите тип участника кнопкой.", PARTICIPANT_TYPES)
                return jsonify({"status": "ok"})

            state["participant_type"] = text

            if text == "Субъект РФ":
                state["participant_name"] = ""
                state["step"] = "region"
                send_message(chat_id, "Введите регион проведения:")
                return jsonify({"status": "ok"})

            state["step"] = "participant_name"

            name_prompts = {
                "ФОИВ": "Укажите полное наименование федерального органа исполнительной власти.",
                "ВУЗ": "Укажите полное наименование образовательной организации высшего образования.",
                "СУЗ": "Укажите полное наименование среднего профессионального образовательного учреждения.",
                "Школа": "Укажите полное наименование общеобразовательной организации.",
                "Организации и НКО": "Укажите полное наименование организации или НКО.",
                "Политические НКО": "Укажите полное наименование политической некоммерческой организации."
            }

            send_message(chat_id, name_prompts.get(text, "Укажите полное наименование участника."))
            return jsonify({"status": "ok"})

        if state["step"] == "participant_name":
            if not text:
                send_message(chat_id, "Укажите наименование участника текстом.")
                return jsonify({"status": "ok"})

            state["participant_name"] = text
            state["step"] = "region"
            send_message(chat_id, "Введите регион проведения:")
            return jsonify({"status": "ok"})

        if state["step"] == "region":
            if not text:
                send_message(chat_id, "Введите регион проведения текстом.")
                return jsonify({"status": "ok"})

            state["region"] = text
            state["step"] = "action"
            send_message(chat_id, "Выберите Всероссийскую акцию:", ACTIONS)
            return jsonify({"status": "ok"})

        if state["step"] == "action":
            if text not in ACTIONS:
                send_message(chat_id, "Пожалуйста, выберите Всероссийскую акцию кнопкой.", ACTIONS)
                return jsonify({"status": "ok"})

            state["action"] = text
            state["step"] = "format"
            send_message(chat_id, "Введите формат мероприятия:")
            return jsonify({"status": "ok"})

        if state["step"] == "format":
            if not text:
                send_message(chat_id, "Введите формат мероприятия текстом.")
                return jsonify({"status": "ok"})

            state["format"] = text
            state["step"] = "disk_link"
            send_message(
                chat_id,
                "Пришлите ссылку на Яндекс.Диск с фото- и видеоматериалами "
                "по данному формату."
            )
            return jsonify({"status": "ok"})

        if state["step"] == "disk_link":
            if not has_link(text):
                send_message(chat_id, "Пожалуйста, пришлите ссылку на Яндекс.Диск с фото- и видеоматериалами.")
                return jsonify({"status": "ok"})

            state["disk_link"] = text

            try:
                save_application_to_excel(state)
            except Exception as excel_error:
                print("EXCEL ERROR:", excel_error, flush=True)
                send_message(
                    chat_id,
                    "Заявка получена, но не удалось записать её в таблицу. "
                    "Пожалуйста, сообщите организаторам."
                )
                return jsonify({"status": "ok"})

            send_message(
                chat_id,
                "Заявка успешно зарегистрирована ✅\n\n"
                f"Тип участника: {state.get('participant_type')}\n"
                f"Наименование участника: {state.get('participant_name') or '—'}\n"
                f"Регион: {state.get('region')}\n"
                f"Всероссийская акция: {state.get('action')}\n"
                f"Формат мероприятия: {state.get('format')}\n"
                f"Ссылка на материалы: {state.get('disk_link')}\n\n"
                "Если у Вас есть материалы по другому формату в рамках этой же акции, "
                "Вы можете добавить ещё одну заявку.\n\n"
                "Выберите дальнейшее действие:",
                ["Добавить ещё один формат", "Завершить работу"]
            )
            return jsonify({"status": "ok"})

    except Exception as e:
        print("ERROR:", e, flush=True)

    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run()
